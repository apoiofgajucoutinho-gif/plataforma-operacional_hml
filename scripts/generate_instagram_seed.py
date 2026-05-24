from __future__ import annotations

import json
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "Insight.xlsx"
OUTPUT_PATH = ROOT / "supabase" / "seed" / "instagram_insight_import.sql"


def sql_string(value: Any) -> str:
    if value is None:
        return "null"
    text = str(value)
    return "'" + text.replace("'", "''") + "'"


def sql_json(value: dict[str, Any]) -> str:
    return sql_string(json.dumps(value, ensure_ascii=False, default=str)) + "::jsonb"


def number_or_zero(value: Any) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def nullable_int(value: Any) -> str:
    if value is None or value == "":
        return "null"
    try:
        return str(int(float(value)))
    except (TypeError, ValueError):
        return "null"


def normalize_post_id(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, float):
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    try:
        decimal = Decimal(text)
        if decimal == decimal.to_integral_value():
            return str(decimal.quantize(Decimal("1")))
    except (InvalidOperation, ValueError):
        pass
    return text


def normalize_type(value: Any) -> str:
    text = str(value or "").lower()
    if "reel" in text or "video" in text:
        return "Reels"
    if "carousel" in text or "carross" in text or "carros" in text or "album" in text:
        return "Carrossel"
    if "image" in text or "foto" in text or "static" in text or "estat" in text:
        return "Estatico"
    return "Outro"


def date_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    if not text:
        raise ValueError("data_postagem vazia")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"data_postagem invalida: {text}")


def timestamp_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    text = str(value or "").strip()
    if not text:
        return ""
    return text


def time_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0).isoformat()
    if isinstance(value, time):
        return value.replace(microsecond=0).isoformat()
    text = str(value or "").strip()
    return text


def engagement(likes: int, comments: int, saved: int | None, reach: int | None) -> tuple[str, str]:
    if not reach or reach <= 0 or (likes <= 0 and comments <= 0):
        return "null", "N/A"
    rate = (likes + comments + (saved or 0)) / reach
    if rate >= 0.05:
        classification = "Bom"
    elif rate >= 0.02:
        classification = "Medio"
    else:
        classification = "Ruim"
    return f"{rate:.6f}", classification


def main() -> None:
    workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["Posts"]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    header_index = {str(header): index + 1 for index, header in enumerate(headers)}
    required = [
        "data_coleta",
        "post_id",
        "tipo_postagem",
        "likes",
        "comentarios",
        "data_postagem",
        "hora_postagem",
        "legenda",
        "permalink",
        "reach",
        "saved",
        "shares",
    ]
    missing = [header for header in required if header not in header_index]
    if missing:
        raise RuntimeError(f"Colunas ausentes: {missing}")

    statements: list[str] = [
        "begin;",
        "",
        "with existing as (",
        "  select id from public.tenants where nome = 'Juliana Coutinho' limit 1",
        "), inserted as (",
        "  insert into public.tenants (nome, tipo)",
        "  select 'Juliana Coutinho', 'cliente'",
        "  where not exists (select 1 from existing)",
        "  returning id",
        "), tenant as (",
        "  select id from existing union all select id from inserted limit 1",
        ")",
        "insert into public.instagram_accounts (tenant_id, nome, username)",
        "select id, 'Juliana Coutinho', 'fga.jucoutinho' from tenant",
        "on conflict (tenant_id, username) do update set nome = excluded.nome;",
        "",
    ]

    total = 0
    failed = 0
    for row in range(2, sheet.max_row + 1):
        raw = {header: sheet.cell(row, header_index[header]).value for header in required}
        if not raw.get("data_postagem") or not raw.get("permalink"):
            failed += 1
            continue

        total += 1
        post_id = normalize_post_id(raw["post_id"])
        likes = number_or_zero(raw["likes"])
        comments = number_or_zero(raw["comentarios"])
        reach = None if raw["reach"] in (None, "") else number_or_zero(raw["reach"])
        saved = None if raw["saved"] in (None, "") else number_or_zero(raw["saved"])
        shares = None if raw["shares"] in (None, "") else number_or_zero(raw["shares"])
        engagement_score, engagement_classification = engagement(likes, comments, saved, reach)
        post_type = normalize_type(raw["tipo_postagem"])

        statements.append(
            "\n".join(
                [
                    "with tenant as (",
                    "  select id from public.tenants where nome = 'Juliana Coutinho' limit 1",
                    "), account as (",
                    "  select ia.id, ia.tenant_id",
                    "  from public.instagram_accounts ia",
                    "  join tenant t on t.id = ia.tenant_id",
                    "  where ia.username = 'fga.jucoutinho'",
                    "  limit 1",
                    "), post_upsert as (",
                    "  insert into public.instagram_posts (",
                    "    tenant_id, account_id, post_id, data_coleta, data_postagem, hora_postagem,",
                    "    tipo_original, tipo, legenda, permalink, raw_payload",
                    "  )",
                    "  select",
                    "    account.tenant_id,",
                    "    account.id,",
                    f"    {sql_string(post_id)},",
                    f"    {sql_string(timestamp_value(raw['data_coleta']))}::timestamptz,",
                    f"    {sql_string(date_value(raw['data_postagem']))}::date,",
                    f"    {sql_string(time_value(raw['hora_postagem']))}::time,",
                    f"    {sql_string(raw['tipo_postagem'])},",
                    f"    {sql_string(post_type)}::public.instagram_post_type,",
                    f"    {sql_string(raw['legenda'])},",
                    f"    {sql_string(raw['permalink'])},",
                    f"    {sql_json(raw)}",
                    "  from account",
                    "  on conflict (tenant_id, post_id) do update set",
                    "    data_coleta = excluded.data_coleta,",
                    "    data_postagem = excluded.data_postagem,",
                    "    hora_postagem = excluded.hora_postagem,",
                    "    tipo_original = excluded.tipo_original,",
                    "    tipo = excluded.tipo,",
                    "    legenda = excluded.legenda,",
                    "    permalink = excluded.permalink,",
                    "    raw_payload = excluded.raw_payload",
                    "  returning id, tenant_id, account_id",
                    ")",
                    "insert into public.instagram_metrics (",
                    "  tenant_id, account_id, post_id, likes, comentarios, alcance, salvos,",
                    "  compartilhamentos, engajamento_score, engajamento_classificacao, origem, raw_payload",
                    ")",
                    "select",
                    "  tenant_id,",
                    "  account_id,",
                    "  id,",
                    f"  {likes},",
                    f"  {comments},",
                    f"  {nullable_int(raw['reach'])},",
                    f"  {nullable_int(raw['saved'])},",
                    f"  {nullable_int(raw['shares'])},",
                    f"  {engagement_score},",
                    f"  {sql_string(engagement_classification)}::public.instagram_engagement_classification,",
                    "  'xlsx_import',",
                    f"  {sql_json(raw)}",
                    "from post_upsert",
                    "on conflict (tenant_id, post_id, origem) do update set",
                    "  likes = excluded.likes,",
                    "  comentarios = excluded.comentarios,",
                    "  alcance = excluded.alcance,",
                    "  salvos = excluded.salvos,",
                    "  compartilhamentos = excluded.compartilhamentos,",
                    "  engajamento_score = excluded.engajamento_score,",
                    "  engajamento_classificacao = excluded.engajamento_classificacao,",
                    "  imported_at = now(),",
                    "  raw_payload = excluded.raw_payload;",
                ]
            )
        )

    statements.extend(
        [
            "",
            "with tenant as (",
            "  select id from public.tenants where nome = 'Juliana Coutinho' limit 1",
            "), account as (",
            "  select ia.id, ia.tenant_id",
            "  from public.instagram_accounts ia",
            "  join tenant t on t.id = ia.tenant_id",
            "  where ia.username = 'fga.jucoutinho'",
            "  limit 1",
            ")",
            "insert into public.instagram_import_runs (",
            "  tenant_id, account_id, source, status, total_rows, inserted_rows, updated_rows, failed_rows, finished_at",
            ")",
            f"select tenant_id, id, 'Insight.xlsx', 'completed', {total + failed}, {total}, 0, {failed}, now() from account;",
            "",
            "commit;",
        ]
    )

    OUTPUT_PATH.write_text("\n\n".join(statements), encoding="utf-8")
    print(json.dumps({"output": str(OUTPUT_PATH), "rows": total, "failed": failed}, ensure_ascii=False))


if __name__ == "__main__":
    main()
