from __future__ import annotations

import hashlib
import json
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r"D:\Jass_\Download_HD\TESTE Fluxo de caixa V2.xlsx")
MIGRATIONS = ROOT / "supabase" / "migrations"
SEED_OUT = MIGRATIONS / "0010_financeiro_seed.sql"
IMPORT_START = 11

MONTHS = {
    "janeiro": 1,
    "fevereiro": 2,
    "marco": 3,
    "março": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
}

COURSES = [
    "Formação AASI Premium",
    "Combo Fono Premium",
    "Imersão Zumbido",
    "Imersão CROS e BiCROS - Indicação, Adaptação, Verificação e Validação",
    "Mapa do Raciocínio Clínico",
    "Protocolo Fono Premium 2026",
    "Pack de modelo zumbido: anamnese, avaliação, orientação e atividades",
    "Fundamentos da Timpanometria",
    "O que considerar antes de precificar sua consulta de zumbido",
    "Manual da acufenometria completa",
    "Passo a passo mascaramento plat",
    "Histórico (a reclassificar)",
]

CENTERS = [
    ("Infoproduto", True),
    ("Clínica", True),
    ("Palestras", True),
    ("Administrativo fixo", True),
    ("Não operacional", True),
]

NATURES = ["Fixo", "Variável", "Operacional", "Não operacional"]

CATEGORIES = [
    ("entrada", "Entradas", "Operacional", "receita_bruta"),
    ("entrada", "Movimentações Financeiras", "Não operacional", "resultado_financeiro"),
    ("saida", "Despesas Administrativas", "Fixo", "despesas_administrativas"),
    ("saida", "Despesas Clínica", "Variável", "despesas_operacionais"),
    ("saida", "Despesas Operacionais", "Operacional", "despesas_operacionais"),
    ("saida", "Despesas Pessoal", "Fixo", "despesas_pessoal"),
    ("saida", "Movimentações Financeiras", "Não operacional", "resultado_financeiro"),
    ("saida", "Deduções da Receita", "Variável", "deducoes"),
    ("saida", "Custos Diretos", "Variável", "custos_diretos"),
    ("saida", "Taxas de Plataforma", "Variável", "taxas_plataforma"),
    ("saida", "Agência de Marketing", "Variável", "coproducao"),
    ("saida", "Comissões de Afiliados", "Variável", "comissoes_afiliados"),
    ("saida", "Outros Custos Diretos", "Variável", "outros_custos_diretos"),
    ("saida", "Depreciação", "Não operacional", "depreciacao"),
    ("saida", "IRPJ/CSLL", "Variável", "irpj_csll"),
]

SUBCATEGORIES = [
    ("entrada", "Entradas", "Clínica", "receita_bruta"),
    ("entrada", "Entradas", "Entradas", "receita_bruta"),
    ("entrada", "Movimentações Financeiras", "Movimentações Financeiras", "resultado_financeiro"),
    ("saida", "Movimentações Financeiras", "Movimentações Financeiras", "resultado_financeiro"),
    ("saida", "Movimentações Financeiras", "Distribuição dos Lucros", "nao_dre"),
    ("saida", "Despesas Administrativas", "Contabilidade", "despesas_administrativas"),
    ("saida", "Despesas Administrativas", "Contas fixas", "despesas_administrativas"),
    ("saida", "Despesas Administrativas", "Equipamentos e materiais", "despesas_administrativas"),
    ("saida", "Despesas Clínica", "Clínica", "despesas_operacionais"),
    ("saida", "Despesas Operacionais", "Equipamentos e materiais", "despesas_operacionais"),
    ("saida", "Despesas Operacionais", "Imposto", "deducoes"),
    ("saida", "Despesas Operacionais", "Pedagógico", "despesas_operacionais"),
    ("saida", "Despesas Operacionais", "Softwares e ferramentas", "despesas_operacionais"),
    ("saida", "Despesas Pessoal", "INSS", "despesas_pessoal"),
    ("saida", "Despesas Pessoal", "Prestação de serviços", "despesas_pessoal"),
    ("saida", "Despesas Pessoal", "Pro-labore", "despesas_pessoal"),
    ("saida", "Despesas Pessoal", "Contabilidade", "despesas_pessoal"),
    ("saida", "Deduções da Receita", "Impostos sobre vendas", "deducoes"),
    ("saida", "Taxas de Plataforma", "Greenn", "taxas_plataforma"),
    ("saida", "Taxas de Plataforma", "Hotmart", "taxas_plataforma"),
    ("saida", "Taxas de Plataforma", "TMB", "taxas_plataforma"),
    ("saida", "Agência de Marketing", "Coprodução", "coproducao"),
    ("saida", "Depreciação", "Depreciação", "depreciacao"),
]


def clean(value: Any) -> str:
    text = "" if value is None else str(value)
    text = " ".join(text.replace("\u00a0", " ").split())
    return text.strip()


def sql(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, (dict, list)):
        return "'" + json.dumps(value, ensure_ascii=False).replace("'", "''") + "'::jsonb"
    return "'" + str(value).replace("'", "''") + "'"


def to_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raise ValueError(f"Unexpected date value: {value!r}")


def competence_month(raw: Any, paid_at: date) -> date:
    if raw is None or clean(raw) == "":
        return date(paid_at.year, paid_at.month, 1)
    if isinstance(raw, datetime):
        return date(raw.year, raw.month, 1)
    if isinstance(raw, date):
        return date(raw.year, raw.month, 1)
    month = MONTHS.get(clean(raw).lower())
    if not month:
        raise ValueError(f"Unknown competence month: {raw!r}")
    year = paid_at.year if month <= paid_at.month else paid_at.year - 1
    return date(year, month, 1)


def normalize_center(value: str) -> str:
    if value.lower() == "clinica":
        return "Clínica"
    if value.lower() in {"nao operacional", "não operacional"}:
        return "Não operacional"
    return value


def normalize_subcategory(value: str) -> str:
    if value.lower() == "distruibuição dos lucros":
        return "Distribuição dos Lucros"
    return value


def tipo_enum(value: str) -> str:
    lowered = value.lower()
    if lowered.startswith("entrada"):
        return "entrada"
    if lowered.startswith("sa"):
        return "saida"
    raise ValueError(f"Unknown tipo: {value}")


def payment_enum(value: str) -> str:
    lowered = value.lower()
    if "cart" in lowered:
        return "cartao_credito"
    if "conta" in lowered:
        return "conta_bancaria"
    if "pix" in lowered:
        return "pix"
    if "boleto" in lowered:
        return "boleto"
    if "dinheiro" in lowered:
        return "dinheiro"
    raise ValueError(f"Unknown payment: {value}")


def money(value: Any) -> str:
    return str(Decimal(str(value)).quantize(Decimal("0.01")))


def row_hash(row: dict[str, Any]) -> str:
    payload = json.dumps(row, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()


def load_rows() -> list[dict[str, Any]]:
    wb = load_workbook(SOURCE, data_only=True, read_only=True)
    ws = wb["Fluxo de caixa V2"]
    headers = [clean(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = dict(zip(headers, values))
        if not record.get("Data"):
            continue
        paid_at = to_date(record["Data"])
        tipo = tipo_enum(clean(record["Tipo"]))
        center = normalize_center(clean(record["Centro de Resultado"]))
        category = clean(record["Categoria"])
        subcategory = normalize_subcategory(clean(record["Subcategoria"]))
        form = payment_enum(clean(record["Pagamento"]))
        compact = {
            "source_row": excel_row,
            "data_pagamento": paid_at.isoformat(),
            "mes_competencia": competence_month(record.get("Mês de competência"), paid_at).isoformat(),
            "centro": center,
            "tipo": tipo,
            "natureza": clean(record["Natureza"]),
            "categoria": category,
            "subcategoria": subcategory,
            "descricao": clean(record["Descrição"]),
            "forma_pagamento": form,
            "valor": money(record["Valor"]),
            "raw_pagamento": clean(record["Pagamento"]),
            "raw_mes_competencia": clean(record.get("Mês de competência")),
        }
        compact["import_hash"] = row_hash(compact)
        rows.append(compact)
    return rows


def seed_values() -> str:
    lines: list[str] = []
    lines.append("-- Auto-generated by scripts/generate_financeiro_seed.py")
    lines.append("-- Source: D:\\Jass_\\Download_HD\\TESTE Fluxo de caixa V2.xlsx")
    lines.append("")
    lines.append("begin;")
    lines.append("")
    lines.append(
        """
create or replace function app_private.fin_auto_provisao_imposto()
returns trigger
language plpgsql
security definer
set search_path = public, app_private
as $$
declare
  v_centro text;
  v_categoria_id uuid;
  v_subcategoria_id uuid;
  v_percentual numeric(7,4);
  v_valor numeric(14,2);
begin
  if new.tipo <> 'entrada'
     or new.status <> 'realizado'
     or new.origem in ('provisao_imposto', 'importacao') then
    return new;
  end if;

  select nome into v_centro
  from public.fin_centros_resultado
  where id = new.centro_resultado_id;

  if v_centro <> 'Infoproduto' then
    return new;
  end if;

  v_percentual := coalesce(
    nullif((select valor->>'percentual_provisao_imposto'
            from public.fin_config
            where tenant_id = new.tenant_id and chave = 'financeiro'), '')::numeric,
    0.06
  );
  v_valor := round(new.valor * v_percentual, 2);

  select c.id, s.id
    into v_categoria_id, v_subcategoria_id
  from public.fin_categorias c
  left join public.fin_subcategorias s
    on s.categoria_id = c.id
   and s.nome = 'Impostos sobre vendas'
  where c.tenant_id = new.tenant_id
    and c.nome = 'Deduções da Receita'
    and c.tipo = 'saida'
  limit 1;

  if v_categoria_id is null or v_subcategoria_id is null or v_valor <= 0 then
    return new;
  end if;

  if not exists (
    select 1
    from public.fin_lancamentos
    where tenant_id = new.tenant_id
      and origem = 'provisao_imposto'
      and lancamento_origem_id = new.id
  ) then
    insert into public.fin_lancamentos (
      tenant_id, data_pagamento, mes_competencia, tipo, status,
      centro_resultado_id, categoria_id, subcategoria_id, forma_pagamento,
      banco_id, descricao, valor, observacao, origem, lancamento_origem_id, created_by
    )
    values (
      new.tenant_id,
      app_private.fin_due_date(date_trunc('month', new.data_pagamento)::date + interval '1 month', 20),
      new.mes_competencia,
      'saida',
      'previsto',
      new.centro_resultado_id,
      v_categoria_id,
      v_subcategoria_id,
      'boleto',
      new.banco_id,
      'Provisão automática de imposto - ' || new.descricao,
      v_valor,
      'Gerado automaticamente a partir do lançamento ' || new.id,
      'provisao_imposto',
      new.id,
      new.created_by
    );
  end if;

  return new;
end;
$$;
""".strip()
    )
    center_values = ",\n    ".join(f"({sql(name)}, {'true' if active else 'false'})" for name, active in CENTERS)
    lines.append(f"""
with tenant_scope as (select id as tenant_id from public.tenants where lower(nome) like '%juliana coutinho%' order by created_at limit 1),
vals(nome, ativo) as (values
    {center_values}
)
insert into public.fin_centros_resultado (tenant_id, nome, ativo)
select t.tenant_id, v.nome, v.ativo from tenant_scope t cross join vals v
on conflict (tenant_id, nome) do update set ativo = excluded.ativo;
""".strip())
    nature_values = ",\n    ".join(f"({sql(name)})" for name in NATURES)
    lines.append(f"""
with tenant_scope as (select id as tenant_id from public.tenants where lower(nome) like '%juliana coutinho%' order by created_at limit 1),
vals(nome) as (values
    {nature_values}
)
insert into public.fin_naturezas (tenant_id, nome, ativo)
select t.tenant_id, v.nome, true from tenant_scope t cross join vals v
on conflict (tenant_id, nome) do update set ativo = excluded.ativo;
""".strip())
    lines.append(
        """
with tenant_scope as (select id as tenant_id from public.tenants where lower(nome) like '%juliana coutinho%' order by created_at limit 1)
insert into public.fin_bancos (tenant_id, nome, apelido, saldo_inicial, ativo)
select tenant_id, 'Histórico (a reclassificar)', 'Histórico', 0, true from tenant_scope
on conflict (tenant_id, nome) do update
set apelido = excluded.apelido, ativo = true;
""".strip()
    )
    lines.append(
        """
with tenant_scope as (select id as tenant_id from public.tenants where lower(nome) like '%juliana coutinho%' order by created_at limit 1),
bank_scope as (
  select b.id as banco_id, b.tenant_id
  from public.fin_bancos b
  join tenant_scope t on t.tenant_id = b.tenant_id
  where b.nome = 'Histórico (a reclassificar)'
)
insert into public.fin_cartoes (tenant_id, nome, banco_id, dia_fechamento, dia_vencimento, limite, ativo)
select tenant_id, 'Histórico (a reclassificar)', banco_id, 3, 10, null, true from bank_scope
on conflict (tenant_id, nome) do update
set banco_id = excluded.banco_id, dia_fechamento = 3, dia_vencimento = 10, ativo = true;
""".strip()
    )
    category_values = ",\n    ".join(
        f"({sql(tipo)}::public.fin_tipo_lancamento, {sql(category)}, {sql(nature)}, {sql(group)})"
        for tipo, category, nature, group in CATEGORIES
    )
    lines.append(f"""
with tenant_scope as (select id as tenant_id from public.tenants where lower(nome) like '%juliana coutinho%' order by created_at limit 1),
vals(tipo, nome, natureza, dre_grupo) as (values
    {category_values}
)
insert into public.fin_categorias (tenant_id, natureza_id, tipo, nome, dre_grupo, ativo)
select t.tenant_id, n.id, v.tipo, v.nome, v.dre_grupo, true
from tenant_scope t
join vals v on true
join public.fin_naturezas n on n.tenant_id = t.tenant_id and n.nome = v.natureza
on conflict (tenant_id, tipo, nome) do update
set natureza_id = excluded.natureza_id, dre_grupo = excluded.dre_grupo, ativo = true;
""".strip())
    subcategory_values = ",\n    ".join(
        f"({sql(tipo)}::public.fin_tipo_lancamento, {sql(category)}, {sql(subcategory)}, {sql(group)})"
        for tipo, category, subcategory, group in SUBCATEGORIES
    )
    lines.append(f"""
with tenant_scope as (select id as tenant_id from public.tenants where lower(nome) like '%juliana coutinho%' order by created_at limit 1),
vals(tipo, categoria, nome, dre_grupo) as (values
    {subcategory_values}
)
insert into public.fin_subcategorias (tenant_id, categoria_id, nome, dre_grupo, ativo)
select t.tenant_id, c.id, v.nome, v.dre_grupo, true
from tenant_scope t
join vals v on true
join public.fin_categorias c on c.tenant_id = t.tenant_id and c.tipo = v.tipo and c.nome = v.categoria
on conflict (tenant_id, categoria_id, nome) do update
set dre_grupo = excluded.dre_grupo, ativo = true;
""".strip())
    course_values = ",\n    ".join(f"({sql(course)})" for course in COURSES)
    lines.append(f"""
with tenant_scope as (select id as tenant_id from public.tenants where lower(nome) like '%juliana coutinho%' order by created_at limit 1),
vals(nome) as (values
    {course_values}
)
insert into public.fin_cursos (tenant_id, nome, ativo)
select t.tenant_id, v.nome, true from tenant_scope t cross join vals v
on conflict (tenant_id, nome) do update set ativo = true;
""".strip())
    lines.append(
        """
with tenant_scope as (select id as tenant_id from public.tenants where lower(nome) like '%juliana coutinho%' order by created_at limit 1)
insert into public.fin_config (tenant_id, chave, valor)
select tenant_id, 'financeiro', '{"limite_capex":5000,"tratamento_coproducao":"despesa","percentual_provisao_imposto":0.06}'::jsonb
from tenant_scope
on conflict (tenant_id, chave) do update
set valor = public.fin_config.valor || excluded.valor;
""".strip()
    )
    lines.append(
        """
with tenant_scope as (select id as tenant_id from public.tenants where lower(nome) like '%juliana coutinho%' order by created_at limit 1),
infoproduto as (
  select id as centro_id, tenant_id
  from public.fin_centros_resultado
  where nome = 'Infoproduto'
),
members as (
  select tm.tenant_id, tm.user_id, tm.role::text as role
  from public.tenant_members tm
  join tenant_scope t on t.tenant_id = tm.tenant_id
)
insert into public.fin_perfis_usuario (tenant_id, user_id, perfil, centros_permitidos, ativo)
select
  m.tenant_id,
  m.user_id,
  case when m.role = 'MARKETING_PARTNER' then 'marketing'::public.fin_perfil_acesso else 'admin'::public.fin_perfil_acesso end,
  case when m.role = 'MARKETING_PARTNER' then array[i.centro_id]::uuid[] else array[]::uuid[] end,
  true
from members m
left join infoproduto i on i.tenant_id = m.tenant_id
where m.role in ('ADMIN', 'MARKETING_PARTNER')
on conflict (tenant_id, user_id) do update
set perfil = excluded.perfil,
    centros_permitidos = excluded.centros_permitidos,
    ativo = true;
""".strip()
    )
    lines.append(
        """
create unique index if not exists fin_lancamentos_import_hash_idx
on public.fin_lancamentos (tenant_id, ((metadata->>'import_hash')))
where origem = 'importacao' and metadata ? 'import_hash';
""".strip()
    )
    return "\n\n".join(lines)


def import_values(rows: list[dict[str, Any]]) -> str:
    import_rows = ",\n    ".join(
        "("
        + ", ".join(
            [
                str(row["source_row"]),
                f"{sql(row['data_pagamento'])}::date",
                f"{sql(row['mes_competencia'])}::date",
                sql(row["centro"]),
                f"{sql(row['tipo'])}::public.fin_tipo_lancamento",
                sql(row["natureza"]),
                sql(row["categoria"]),
                sql(row["subcategoria"]),
                sql(row["descricao"]),
                f"{sql(row['forma_pagamento'])}::public.fin_forma_pagamento",
                f"{row['valor']}::numeric",
                sql(row["raw_pagamento"]),
                sql(row["raw_mes_competencia"]),
                sql(row["import_hash"]),
            ]
        )
        + ")"
        for row in rows
    )
    return (
        f"""
create temporary table if not exists fin_import_raw (
  source_row integer,
  data_pagamento date,
  mes_competencia date,
  centro text,
  tipo public.fin_tipo_lancamento,
  natureza text,
  categoria text,
  subcategoria text,
  descricao text,
  forma_pagamento public.fin_forma_pagamento,
  valor numeric(14,2),
  raw_pagamento text,
  raw_mes_competencia text,
  import_hash text
) on commit drop;

truncate table fin_import_raw;

insert into fin_import_raw (
  source_row, data_pagamento, mes_competencia, centro, tipo, natureza,
  categoria, subcategoria, descricao, forma_pagamento, valor,
  raw_pagamento, raw_mes_competencia, import_hash
)
values
    {import_rows};

do $$
declare
  v_missing integer;
begin
  with tenant_scope as (
    select id as tenant_id
    from public.tenants
    where lower(nome) like '%juliana coutinho%'
    order by created_at
    limit 1
  ),
  checked as (
    select r.source_row
    from fin_import_raw r
    cross join tenant_scope t
    left join public.fin_centros_resultado cr
      on cr.tenant_id = t.tenant_id and cr.nome = r.centro
    left join public.fin_categorias c
      on c.tenant_id = t.tenant_id and c.tipo = r.tipo and c.nome = r.categoria
    left join public.fin_subcategorias s
      on s.tenant_id = t.tenant_id and s.categoria_id = c.id and s.nome = r.subcategoria
    left join public.fin_bancos b
      on b.tenant_id = t.tenant_id and b.nome = 'Histórico (a reclassificar)'
    left join public.fin_cartoes ca
      on ca.tenant_id = t.tenant_id and ca.nome = 'Histórico (a reclassificar)'
    left join public.fin_cursos cu
      on cu.tenant_id = t.tenant_id and cu.nome = 'Histórico (a reclassificar)'
    where cr.id is null
       or c.id is null
       or s.id is null
       or (r.forma_pagamento <> 'cartao_credito' and b.id is null)
       or (r.forma_pagamento = 'cartao_credito' and ca.id is null)
       or (r.centro = 'Infoproduto' and r.tipo = 'entrada' and cu.id is null)
  )
  select count(*) into v_missing from checked;

  if v_missing > 0 then
    raise exception 'Financeiro importacao: % linhas sem cadastro de referencia', v_missing;
  end if;
end;
$$;

with tenant_scope as (
  select id as tenant_id
  from public.tenants
  where lower(nome) like '%juliana coutinho%'
  order by created_at
  limit 1
),
refs as (
  select
    r.*,
    t.tenant_id,
    cr.id as centro_resultado_id,
    c.id as categoria_id,
    s.id as subcategoria_id,
    b.id as banco_id,
    ca.id as cartao_id,
    cu.id as curso_id
  from fin_import_raw r
  cross join tenant_scope t
  join public.fin_centros_resultado cr
    on cr.tenant_id = t.tenant_id and cr.nome = r.centro
  join public.fin_categorias c
    on c.tenant_id = t.tenant_id and c.tipo = r.tipo and c.nome = r.categoria
  join public.fin_subcategorias s
    on s.tenant_id = t.tenant_id and s.categoria_id = c.id and s.nome = r.subcategoria
  left join public.fin_bancos b
    on b.tenant_id = t.tenant_id and b.nome = 'Histórico (a reclassificar)'
  left join public.fin_cartoes ca
    on ca.tenant_id = t.tenant_id and ca.nome = 'Histórico (a reclassificar)'
  left join public.fin_cursos cu
    on cu.tenant_id = t.tenant_id and cu.nome = 'Histórico (a reclassificar)'
)
insert into public.fin_lancamentos (
  tenant_id, data_pagamento, mes_competencia, tipo, status,
  centro_resultado_id, categoria_id, subcategoria_id, curso_id,
  forma_pagamento, banco_id, cartao_id, qtd_parcelas, parcela_numero,
  descricao, valor, observacao, origem, status_origem, metadata
)
select
  tenant_id,
  data_pagamento,
  mes_competencia,
  tipo,
  'realizado'::public.fin_status_lancamento,
  centro_resultado_id,
  categoria_id,
  subcategoria_id,
  case when centro = 'Infoproduto' and tipo = 'entrada' then curso_id else null::uuid end,
  forma_pagamento,
  case when forma_pagamento = 'cartao_credito' then null::uuid else banco_id end,
  case when forma_pagamento = 'cartao_credito' then cartao_id else null::uuid end,
  1,
  1,
  descricao,
  valor,
  'Importado da planilha Fluxo de caixa V2',
  'importacao'::public.fin_origem_lancamento,
  'planilha_v2',
  jsonb_build_object(
    'source', 'TESTE Fluxo de caixa V2.xlsx',
    'source_row', source_row,
    'import_hash', import_hash,
    'raw_pagamento', raw_pagamento,
    'raw_mes_competencia', raw_mes_competencia
  )
from refs
where not exists (
  select 1
  from public.fin_lancamentos l
  where l.tenant_id = refs.tenant_id
    and l.origem = 'importacao'
    and l.metadata->>'import_hash' = refs.import_hash
);
""".strip()
    )


def main() -> None:
    rows = load_rows()
    old = MIGRATIONS / "0010_financeiro_seed_import.sql"
    if old.exists():
        old.unlink()
    for old_batch in MIGRATIONS.glob("0011_financeiro_import_lote_*.sql"):
        old_batch.unlink()

    seed_content = seed_values() + "\n\ncommit;\n"
    SEED_OUT.write_text(seed_content, encoding="utf-8", newline="\n")
    print(f"Wrote {SEED_OUT}")

    batch_size = 25
    for idx, start in enumerate(range(0, len(rows), batch_size), start=1):
        chunk = rows[start : start + batch_size]
        out = MIGRATIONS / f"{IMPORT_START + idx - 1:04d}_financeiro_import_lote_{idx:02d}.sql"
        content = "-- Auto-generated by scripts/generate_financeiro_seed.py\n"
        content += f"-- Source rows: {chunk[0]['source_row']}..{chunk[-1]['source_row']}\n\n"
        content += "begin;\n\n"
        content += import_values(chunk)
        content += "\n\ncommit;\n"
        out.write_text(content, encoding="utf-8", newline="\n")
        print(f"Wrote {out}")
    print(f"Rows: {len(rows)}")


if __name__ == "__main__":
    main()
